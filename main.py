import os
import glob
from datetime import date

from file_paths import invnday, metal_master, steel_codes, alum_codes, copper_codes, save, report

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

today = date.today().isoformat().replace("-","")

list_of_files = glob.glob(invnday)
latest_file = max(list_of_files, key=os.path.getctime)

inv_wb = openpyxl.load_workbook(latest_file)
inv_ws = inv_wb['INVNDAY']

metal_master_wb = openpyxl.load_workbook(metal_master)
metal_master_ws = metal_master_wb['Sheet1']

final_wb = Workbook()
final_ws = final_wb.active

has_data = True  # While loop used to find the number of rows in the inv sheet that contain data.
inv_row_count = 0

while has_data:
    inv_row_count += 1
    data = inv_ws.cell(row=inv_row_count, column=1).value
    if data == None:
        has_data = False

has_data = True  # While loop used to find the number of rows in the metal_master sheet that contain data.
metal_master_row_count = 0

while has_data:
    metal_master_row_count += 1
    data = metal_master_ws.cell(row=metal_master_row_count, column=10).value
    if data == None:
        has_data = False

def harm_codes(fp):  # Returns the contents of a text file as a list of strings from a file path.
    codes = []
    with open(fp) as f:
        file_contents = f.read()    
    for content in file_contents.split():
        codes.append(content)
    return codes

def find_declaration_req(codes):  # Finds the codes in workbook that require steel declaration.
    declaration_req = []
    for i in range(1, inv_row_count):    
        for hc in codes:
            value = inv_ws.cell(row=i, column=11).value
            if value[:len(hc)] == hc:
                declaration_req.append(inv_ws.cell(row=i, column=11))
                break
    return declaration_req

def declared_sku(harm_cell):  # Finds sku associated with harm requiring declaration.
    skus = []
    for cell in harm_cell:
        skus.append(inv_ws.cell(row=cell.row, column=4))
    return skus

steel_sku = declared_sku(find_declaration_req(harm_codes(steel_codes)))
alum_sku = declared_sku(find_declaration_req(harm_codes(alum_codes)))
copper_sku = declared_sku(find_declaration_req(harm_codes(copper_codes)))

declared_ranges = []  # Tracking all ranges declared with relevant sku/cell pairing for sorting.

def range_declaration(skus, metal):  # finds the cell ranges in metal_master_ws that need to be declared and adds them to a list.

    for sku in skus:
        needs_declaration = True  # used to track if a specific metal for a sku is declared. Eg, a code needing both aluminum & steel.

        for i in range(1, metal_master_row_count):
            value = metal_master_ws.cell(row=i, column=3).value

            if sku.value == value:
                if metal in metal_master_ws.cell(row=i+2, column=1).value.lower():
                    needs_declaration = False
                    declared_range_and_sku = []
                    
                    metal_master_range = metal_master_ws[f"A{i-1}" : f"I{i+2}"]  # Cell range of information relevant to the sku.
                    metal_master_cells = [val for row in metal_master_range for val in row]  # nested tuple unpacking.

                    declared_range_and_sku.append(metal_master_cells)
                    declared_range_and_sku.append(sku)
                    declared_ranges.append(declared_range_and_sku)
                    
        if needs_declaration == True:
            print(f"Sku {sku.value} on line {sku.row} needs {metal} to be declared.")
            with open(f"{report}/metal_declaration_report{today}", "a") as f:
                f.write(f"Sku {sku.value} needs {metal} to be declared.\n")

def range_sort(ranges):
    ranges_sorted = sorted(ranges, key=lambda pairing: pairing[1].row)
    return ranges_sorted

final_ws_tracker = 0  # Used to keep progress in the final worksheet and not overwrite data.

def final_ws_editing(sorted_ranges):  # Adds shipment info to the final sheet for each sku that needs to be declared.
    global final_ws_tracker

    for list_range in sorted_ranges:
        metal_master_cells = list_range[0]
        sku = list_range[1]

        final_range = final_ws[f"A{1+final_ws_tracker}" : f"I{4+final_ws_tracker}"]  # Cell range to have information added.
        final_cells = [val for row in final_range for val in row]
        
        for i in range(len(final_cells)):
            final_ws[final_cells[i].coordinate] = metal_master_ws[metal_master_cells[i].coordinate].value
        final_ws[f"A{2 + final_ws_tracker}"] = inv_ws[f"A{sku.row}"].value  # Adds in the shipment ID.
        final_ws[f"B{2 + final_ws_tracker}"] = inv_ws[f"C{sku.row}"].value  # Adds in the Invoice Number.
        final_ws[f"D{2 + final_ws_tracker}"] = inv_ws[f"H{sku.row}"].value  # Changes the quanity of items sent.
        final_ws[f"F{2 + final_ws_tracker}"] = round(float(inv_ws[f"J{sku.row}"].value) # float() used to work around ' infront of numbers
                                                    * float(final_ws[f"F{2 + final_ws_tracker}"].value), 2)
        final_ws[f"I{2 + final_ws_tracker}"] = (float(final_ws[f"D{2 + final_ws_tracker}"].value)  # Calculates total value of metal.
                                                * float(final_ws[f"F{2 + final_ws_tracker}"].value))
        
        final_ws_formatting()

        final_ws_tracker += 5  # Increment progress by 5 to leave a space between each declared sku.

def final_ws_formatting():  # Applies bold and borders to sections of excel to keep data visually separate in the final sheet.
    font_bold = Font(bold= True)        
    border = Side(border_style= "thin")
    border_top = Border(top= border)
    border_bottom = Border(bottom= border)

    first_row_range = final_ws[f"A{1 + final_ws_tracker}" : f"I{1 + final_ws_tracker}"]
    first_row = [val for row in first_row_range for val in row]
    
    third_row_range = final_ws[f"A{3 + final_ws_tracker}" : f"I{3 + final_ws_tracker}"]
    third_row = [val for row in third_row_range for val in row]

    fourth_row_range = final_ws[f"A{4 + final_ws_tracker}" : f"I{4 + final_ws_tracker}"]
    fourth_row = [val for row in fourth_row_range for val in row]

    for cell in first_row:
        final_ws[cell.coordinate].font = font_bold
        final_ws[cell.coordinate].border = border_top

    for cell in third_row:
        final_ws[cell.coordinate].font = font_bold

    for cell in fourth_row:
        final_ws[cell.coordinate].border = border_bottom

range_declaration(steel_sku, "steel")
range_declaration(alum_sku, "aluminum")
range_declaration(copper_sku, "copper")
final_ws_editing(range_sort(declared_ranges))

final_wb.save(f"{save}/final_test_{today}.xlsx")

if os.path.isfile(f"{report}/metal_declaration_report{today}"):
    print("",f"\nSkus to be delcared saved in reports folder as metal_declaration_report{today} inside ronelle_close_files on the desktop.")
else:
    print("\nNo additional skus to be declared.\n")
input("Press Enter to exit...")